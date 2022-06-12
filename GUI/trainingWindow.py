import os
import platform

import joblib
import matplotlib
from PyQt6.QtCore import QSize
from PyQt6.QtGui import QPalette, QColor, QPixmap
from PyQt6.QtWidgets import QWidget, QPushButton, QCheckBox, QLabel, QLineEdit, QVBoxLayout, QFileDialog
from openpyxl import load_workbook
from sklearn.metrics import accuracy_score
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import MinMaxScaler
import matplotlib.pyplot as plt
from Database.modelDataGeneration import get_model_data
from GUI.utils import lightgray, pop_message_box
from Model.modelCreation import rf_model
from variables import DIV

########################################################################################################################
# Represents the window that lets the user train their own model
########################################################################################################################

maxExcludedFeatures = 10

class TrainModelWindow(QWidget):
    def __init__(self, dashboard):
        super().__init__()
        self.model = None
        self.excelRowIndex = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
        self.setFixedSize(QSize(700, 800))
        self.setWindowTitle("Train a model")
        # close home page
        self.dashboard = dashboard
        dashboard.close()

        self.displayWindow = None

        # Creating a home button
        self.home_button = QPushButton('Home')

        # Creating elements
        self.wbc = QCheckBox("WBC")
        self.lymf = QCheckBox("LYMF")
        self.gran = QCheckBox("GRAN")
        self.mid = QCheckBox("MID")
        self.rbc = QCheckBox("RBC")
        self.hgb = QCheckBox("HGB")
        self.mch = QCheckBox("MCH")
        self.mchc = QCheckBox("MCHC")
        self.mpv = QCheckBox("MPV")
        self.plt = QCheckBox("PLT")
        self.accu_label = QLabel()
        self.display_feature_importance_graph = QLabel()
        self.feature_importance = QLabel()
        self.train_button = QPushButton('Train')
        self.input_model_name = QLineEdit()
        self.save_button = QPushButton('Save')

        # Setting widget properties:
        self.setLayout(self.set_elements())
        self.setAutoFillBackground(True)
        q = self.palette()
        q.setColor(QPalette.ColorRole.Window, QColor(lightgray))
        self.setPalette(q)

    # re-opens the dashboard and closes the current window
    def go_to_home(self):
        self.dashboard.show()
        self.close()

    def set_elements(self):
        """
        Sets the elements of the window
        :return: the layout of the window
        """
        # Setting elements:
        self.accu_label.setText("Model accuracy")
        self.feature_importance.setText("Feature importance")
        self.input_model_name.setPlaceholderText("Enter model name")
        self.train_button.clicked.connect(self.train_new_model)
        self.save_button.clicked.connect(self.save_model)
        self.home_button.clicked.connect(self.go_to_home)
        # Adding elements to the layout:
        layout = QVBoxLayout()
        layout.addWidget(self.wbc)
        layout.addWidget(self.lymf)
        layout.addWidget(self.gran)
        layout.addWidget(self.mid)
        layout.addWidget(self.rbc)
        layout.addWidget(self.hgb)
        layout.addWidget(self.mch)
        layout.addWidget(self.mchc)
        layout.addWidget(self.mpv)
        layout.addWidget(self.plt)
        layout.addWidget(self.accu_label)
        layout.addWidget(self.feature_importance)
        layout.addSpacing(10)
        layout.addWidget(self.train_button)
        layout.addWidget(self.input_model_name)
        layout.addWidget(self.save_button)
        layout.addWidget(self.home_button)
        layout.addWidget(self.display_feature_importance_graph, 1)
        return layout

    # Saves a user trained model
    def save_model(self):
        model_name = self.input_model_name.text()
        if model_name == "":
            pop_message_box("Please enter a model name.")
        else:
            # the user tries to save a model only after training it
            if self.model is not None:
                import_path = QFileDialog.getExistingDirectoryUrl().path()
                if import_path != "":
                    import_path = import_path + DIV + model_name + '.pkl'
                    # save the model details into the Excel file (featuresChecklist.xlsx)
                    is_successful = save_features(self.excelRowIndex, model_name)
                    if is_successful == 1:
                        if platform.system() == "Windows":
                            import_path = import_path[1:]
                        joblib.dump(self.model, import_path)
                        # pops a message box
                        pop_message_box("Model saved successfully")
                        self.model = None
            else:
                pop_message_box("Please train a model first.")
            self.input_model_name.setText("")

    # Trains a model based on the features selected
    def train_new_model(self):
        dataset_labeled_seals = get_model_data()
        # Features not included in training
        excluded_features_num = 0
        self.excelRowIndex = [2, 3, 4, 5, 6, 7, 8, 9, 10, 11]
        # if a feature is unchecked, it is removed from the training model params
        if not self.wbc.isChecked():
            dataset_labeled_seals = dataset_labeled_seals.drop(['WBC'], axis=1)  # drop tag column
            self.excelRowIndex.remove(2)
            excluded_features_num += 1
        if not self.lymf.isChecked():
            dataset_labeled_seals = dataset_labeled_seals.drop(['LYMF'], axis=1)  # drop tag column
            self.excelRowIndex.remove(3)
            excluded_features_num += 1
        if not self.gran.isChecked():
            dataset_labeled_seals = dataset_labeled_seals.drop(['GRAN'], axis=1)  # drop tag column
            self.excelRowIndex.remove(4)
            excluded_features_num += 1
        if not self.mid.isChecked():
            dataset_labeled_seals = dataset_labeled_seals.drop(['MID'], axis=1)  # drop tag column
            self.excelRowIndex.remove(5)
            excluded_features_num += 1
        if not self.rbc.isChecked():
            dataset_labeled_seals = dataset_labeled_seals.drop(['RBC'], axis=1)  # drop tag column
            self.excelRowIndex.remove(6)
            excluded_features_num += 1
        if not self.hgb.isChecked():
            dataset_labeled_seals = dataset_labeled_seals.drop(['HGB'], axis=1)  # drop tag column
            self.excelRowIndex.remove(7)
            excluded_features_num += 1
        if not self.mch.isChecked():
            dataset_labeled_seals = dataset_labeled_seals.drop(['MCH'], axis=1)  # drop tag column
            self.excelRowIndex.remove(8)
            excluded_features_num += 1
        if not self.mchc.isChecked():
            dataset_labeled_seals = dataset_labeled_seals.drop(['MCHC'], axis=1)  # drop tag column
            self.excelRowIndex.remove(9)
            excluded_features_num += 1
        if not self.mpv.isChecked():
            dataset_labeled_seals = dataset_labeled_seals.drop(['MPV'], axis=1)  # drop tag column
            self.excelRowIndex.remove(10)
            excluded_features_num += 1
        if not self.plt.isChecked():
            dataset_labeled_seals = dataset_labeled_seals.drop(['PLT'], axis=1)  # drop tag column
            self.excelRowIndex.remove(11)
            excluded_features_num += 1

        if excluded_features_num != maxExcludedFeatures:
            random_forest = rf_model()
            X = dataset_labeled_seals.drop(['Survival'], axis=1)  # separate features from labels
            scaler = MinMaxScaler()
            X = scaler.fit_transform(X)  # normalize the data ( MinMaxScaler ) - scale the data to be between 0 and 1
            y = dataset_labeled_seals['Survival'].values  # get all labels

            x_train, x_test, y_train, y_test = train_test_split(X, y, test_size=0.3,
                                                                random_state=42)  # split data into training
            # and test
            self.model = random_forest.fit(x_train, y_train)  # train the model
            predictions = random_forest.predict(x_test)  # make predictions on the test set
            # get and set accuracy
            self.accu_label.setText(
                "Your new model's accuracy " + str(round(accuracy_score(y_test, predictions) * 100, 1)) + "%")
            feature_importance_str = "Feature importance: " + "\n"
            # get and set feature importance
            for i in range(len((dataset_labeled_seals.drop(['Survival'], axis=1)).columns.values)):
                feature_importance_str = feature_importance_str + str((dataset_labeled_seals.drop(['Survival'], axis=1)).columns.values[i]) + ": " + str(round((random_forest.feature_importances_*100)[i], 1)) + "%\n"
            self.feature_importance.setText(feature_importance_str)
            pop_message_box("Model trained successfully")
        else:
            pop_message_box("Please select features to train on.")



# This function is used to update the Excel file. This Excel file is used to note down all the selected features
# for the trained models. If a feature was used, then the corresponding cell will be filled with 1.
def save_features(row_index_list, model_name):
    # load a workbook and worksheet.
    try:
        wb = load_workbook("featuresChecklist.xlsx")
        ws = wb.active
        max_col = ws.max_column
        is_model_name_unique = True
        for j in range(2, max_col + 1):
            if ws.cell(row=1, column=j).value == model_name:
                pop_message_box("Model names need to be unique, please enter a new model name")
                is_model_name_unique = False
                break
        if is_model_name_unique:
            ws.insert_cols(max_col + 1)
            # Fill in the corresponding values
            ws.cell(row=1, column=max_col + 1).value = model_name
            for i in row_index_list:
                # fill in the cell with 1
                ws.cell(row=i, column=max_col + 1).value = 1
            wb.save("featuresChecklist.xlsx")
            return 1
        else:
            return 0
    except LookupError:
        pop_message_box("Can't find the featuresChecklist.xlsx file")
    return 0

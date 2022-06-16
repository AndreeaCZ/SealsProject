from pathlib import Path
import joblib
import platform
import pandas as pd
import numpy as np
from PyQt6.QtCore import QSize, Qt
from PyQt6.QtGui import QPalette, QColor
from PyQt6.QtWidgets import QGridLayout, QWidget, QLineEdit, QPushButton, QLabel, QComboBox, QFileDialog
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from GUI.utils import lightgray, pop_message_box, get_chances_str_from_int, get_seal_species_str_from_int, \
    get_sex_str_from_int, get_sex_int, get_seal_species_int, find_prediction, darkgray, make_prediction
from variables import MODEL_PATH, DIV, FEATURE_CHECKLIST_PATH

########################################################################################################################
# Represents the window where the user can predict the outcome of a seal
########################################################################################################################

defaultFeatureList = ["WBC", "LYMF", "GRAN", "MID", "RBC", "HGB", "MCH", "MCHC", "MPV", "PLT"]
defaultModelName = "defaultModel"


class PredictionWindow(QWidget):
    def __init__(self, dashboard):
        super().__init__()
        self.sealDataImport = None
        self.sealDataManualInput = None
        self.modelName = defaultModelName
        self.featureList = defaultFeatureList
        self.model = joblib.load(MODEL_PATH)
        self.setWindowTitle("Run predictions")
        self.setFixedSize(QSize(600, 700))
        self.dashboard = dashboard
        dashboard.close()

        # Creating elements:
        self.input_filename = QLineEdit()
        self.import_button = QPushButton('Import')
        self.output_label = QLabel('Enter data to see prediction')
        self.save_button = QPushButton('Save')
        self.load_model_button = QPushButton('Load a model')
        self.load_default_model_button = QPushButton('Default model')
        self.info_label_1 = QLabel('Select sex and enter a seal tag')
        self.info_label_2 = QLabel('Select what model to use')
        self.info_label_3 = QLabel('Enter blood test values')
        self.info_label_4 = QLabel('Click to run prediction once all values are entered')
        self.info_label_5 = QLabel('Click to import values from an excel file')
        self.seal_tag_input = QLineEdit()
        self.run_predict_button = QPushButton('Predict')

        # Creating import subfields
        self.combo1 = QComboBox()
        self.wbc_label = QLabel("WBC: ")
        self.lymf_label = QLabel("LYMF: ")
        self.gran_label = QLabel("GRAN: ")
        self.mid_label = QLabel("MID: ")
        self.rbc_label = QLabel("RBC: ")
        self.hgb_label = QLabel("HGB: ")
        self.mch_label = QLabel("MCH: ")
        self.mchc_label = QLabel("MCHC: ")
        self.mpv_label = QLabel("MPV: ")
        self.plt_label = QLabel("PLT: ")
        self.wbc_input = QLineEdit()
        self.lymf_input = QLineEdit()
        self.gran_input = QLineEdit()
        self.mid_input = QLineEdit()
        self.rbc_input = QLineEdit()
        self.hgb_input = QLineEdit()
        self.mch_input = QLineEdit()
        self.mchc_input = QLineEdit()
        self.mpv_input = QLineEdit()
        self.plt_input = QLineEdit()

        # feature to text field dictionary
        default_input_list = [self.wbc_input, self.lymf_input, self.gran_input, self.mid_input, self.rbc_input,
                            self.hgb_input, self.mch_input, self.mchc_input, self.mpv_input, self.plt_input]
        self.featureInputDict = dict(zip(defaultFeatureList, default_input_list))

        # Creating a home button
        self.home_button = QPushButton('Home')

        # Set widget properties:
        self.setLayout(self.set_elements())
        self.setAutoFillBackground(True)
        q = self.palette()
        q.setColor(QPalette.ColorRole.Window, QColor(lightgray))
        self.setPalette(q)

    def go_to_home(self):
        """
        re-opens the dashboard and closes the current window
        """
        self.dashboard.show()
        self.close()

    def set_elements(self):
        """
        Sets the elements of the window
        :return: the layout of the window
        """
        # Setting elements properties:
        self.combo1.addItem("Female")
        self.combo1.addItem("Male")
        self.input_filename.setPlaceholderText('Enter an excel file name')
        self.import_button.clicked.connect(self.get_import)
        self.load_model_button.clicked.connect(self.load_model)
        self.load_default_model_button.clicked.connect(self.load_default_model)
        self.save_button.clicked.connect(self.save_results)
        self.home_button.clicked.connect(self.go_to_home)
        self.info_label_1.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.info_label_2.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.info_label_3.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.info_label_4.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.info_label_5.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.output_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        self.run_predict_button.clicked.connect(self.get_manual_input_prediction)
        self.output_label.setAutoFillBackground(True)
        p = self.output_label.palette()
        p.setColor(QPalette.ColorRole.Window, QColor(darkgray))
        self.output_label.setPalette(p)

        # Adding elements to input layout:
        layout = QGridLayout()
        layout.addWidget(self.info_label_1, 0, 0, 1, 2)
        layout.addWidget(self.info_label_2, 0, 3, 1, 2)
        layout.addWidget(self.combo1, 1, 0, 1, 2)
        layout.addWidget(self.load_default_model_button, 1, 3, 1, 2)
        layout.addWidget(self.seal_tag_input, 2, 0, 1, 2)
        layout.addWidget(self.load_model_button, 2, 3, 1, 2)
        layout.setRowMinimumHeight(3, 50)  # Row 3 is empty and is used as a space
        layout.addWidget(self.info_label_3, 4, 0, 1, 5)

        layout.addWidget(self.wbc_label, 5, 0)
        layout.addWidget(self.lymf_label, 5, 1)
        layout.addWidget(self.gran_label, 5, 2)
        layout.addWidget(self.mid_label, 5, 3)
        layout.addWidget(self.rbc_label, 5, 4)
        layout.addWidget(self.wbc_input, 6, 0)
        layout.addWidget(self.lymf_input, 6, 1)
        layout.addWidget(self.gran_input, 6, 2)
        layout.addWidget(self.mid_input, 6, 3)
        layout.addWidget(self.rbc_input, 6, 4)

        layout.addWidget(self.hgb_label, 7, 0)
        layout.addWidget(self.mch_label, 7, 1)
        layout.addWidget(self.mchc_label, 7, 2)
        layout.addWidget(self.mpv_label, 7, 3)
        layout.addWidget(self.plt_label, 7, 4)
        layout.addWidget(self.hgb_input, 8, 0)
        layout.addWidget(self.mch_input, 8, 1)
        layout.addWidget(self.mchc_input, 8, 2)
        layout.addWidget(self.mpv_input, 8, 3)
        layout.addWidget(self.plt_input, 8, 4)

        layout.addWidget(self.info_label_4, 11, 0, 1, 4)
        layout.addWidget(self.run_predict_button, 11, 4)
        layout.setRowMinimumHeight(12, 50)  # Row 12 is empty and is used as a space
        layout.addWidget(self.info_label_5, 13, 0, 1, 4)
        layout.addWidget(self.import_button, 13, 4)
        layout.addWidget(self.output_label, 14, 0, 1, 5)
        layout.setRowMinimumHeight(14, 120)
        layout.addWidget(self.input_filename, 18, 0, 1, 4)
        layout.addWidget(self.save_button, 18, 4)
        layout.addWidget(self.home_button, 19, 1, 1, 3)
        layout.setRowMinimumHeight(19, 70)
        layout.setRowStretch(20, 1)
        return layout

    def update_input_and_labels(self):
        '''
        update the input fields and labels based on the loaded model
        '''
        for input in self.featureInputDict.values():
            input.setDisabled(False)
            input.setStyleSheet("background-color: white;")
        # show labels and fields of features in the current feature list
        for i in defaultFeatureList:
            if i not in self.featureList:
                self.featureInputDict.get(i).setStyleSheet("background-color: darkgrey;")
                self.featureInputDict.get(i).setDisabled(True)

    def reset_input_fields_text(self):
        '''
        clear all input fields
        '''
        for input in self.featureInputDict.values():
            input.setText("")

    def load_default_model_with_new_msg(self):
        '''
        loads the default model with a different message
        '''
        self.output_label.setText("")
        self.modelName = defaultModelName
        self.model = joblib.load(MODEL_PATH)
        self.featureList = defaultFeatureList
        pop_message_box("Can't find the featuresChecklist.xlsx file. Default model loaded successfully")

    def update_feature_list(self, filename):
        '''
        Update the list of features used for predicting based on the parameters the new model was trained on
        '''
        try:
            wb = load_workbook(FEATURE_CHECKLIST_PATH)
            ws = wb.active
            max_col = ws.max_column
            max_row = ws.max_row
            feature_list_temp = []
            flag = False
            col_num = -1
            # find the column corresponding to the loaded model
            for j in range(2, max_col + 1):
                if ws.cell(row=1, column=j).value == filename:
                    flag = True
                    col_num = j
                    break
            # The model file exists but its data is not present in the features checklist
            if not flag:
                pop_message_box("Please select another model")
            else:
                # create a list of features the model was trained on
                self.modelName = ws.cell(row=1, column=col_num).value
                for i in range(2, max_row + 1):
                    if ws.cell(row=i, column=col_num).value is not None:
                        feature_list_temp.append(ws.cell(row=i, column=1).value)
                self.featureList = feature_list_temp
                pop_message_box("Model loaded successfully")
        except FileNotFoundError:
            self.load_default_model_with_new_msg()

    def save_results(self):
        '''
        main function to save the results of the prediction
        '''
        file_name = self.input_filename.text()
        if file_name == "":
            pop_message_box("Please enter a file name first.")
        else:
            if self.sealDataManualInput is not None or self.sealDataImport is not None:
                import_path = QFileDialog.getExistingDirectoryUrl().path()
                if import_path != "":
                    import_path = import_path + DIV + file_name + '.xlsx'
                    if platform.system() == "Windows":
                        import_path = import_path[1:]
                    self.save_prediction(import_path)
                    # pops a message box
                    pop_message_box("Prediction saved successfully")
                    self.output_label.setText("")
            else:
                pop_message_box("Please predict something first.")
                self.input_filename.setText("")
            self.sealDataImport = None
            self.sealDataManualInput = None
        self.seal_tag_input.setText("")
        self.input_filename.setText("")

    def save_prediction(self, import_path):
        '''
        Saves the results of the prediction
        '''
        # Create an Excel file
        excel_file = Workbook()
        spread_sheet = excel_file.active

        if self.sealDataManualInput is None:
            seal_data = self.sealDataImport
        else:
            seal_data = self.sealDataManualInput

        feature_list_length = len(self.featureList)
        seal_data_length = len(seal_data)

        # Fill in the features
        spread_sheet.cell(row=1, column=1).value = "SEAL-TAG"
        spread_sheet.cell(row=2, column=1).value = "MODEL NAME"
        for i in range(feature_list_length):
            spread_sheet.cell(row=i + 3, column=1).value = self.featureList[i]
        spread_sheet.cell(row=feature_list_length + 3, column=1).value = "SEX"
        spread_sheet.cell(row=feature_list_length + 4, column=1).value = "SURVIVAL"

        # Fill in the values
        spread_sheet.cell(row=1, column=2).value = seal_data[0]
        spread_sheet.cell(row=2, column=2).value = self.modelName
        for i in range(len(seal_data) - 3):
            spread_sheet.cell(row=i + 3, column=2).value = seal_data[i + 1]

        spread_sheet.cell(row=seal_data_length, column=2).value = get_sex_str_from_int(int(float(seal_data[seal_data_length - 2])))
        spread_sheet.cell(row=seal_data_length + 1, column=2).value = get_chances_str_from_int(int(float(
            seal_data[seal_data_length - 1])))

        excel_file.save(import_path)
        self.reset_input_fields_text()

    def load_default_model(self):
        '''
        Loads the default model
        '''
        self.output_label.setText("")
        self.seal_tag_input.setText("")
        self.modelName = defaultModelName
        self.model = joblib.load(MODEL_PATH)
        self.featureList = defaultFeatureList
        pop_message_box("Default model loaded successfully")
        self.reset_input_fields_text()
        self.update_input_and_labels()

    def load_model(self):
        '''
        Loads a model from the local machine
        '''
        self.output_label.setText("")
        self.seal_tag_input.setText("")
        import_path = QFileDialog.getOpenFileName(filter='PKL files (*.pkl)')[0]
        if import_path != "":
            self.model = joblib.load(import_path)
            file_name = Path(import_path).stem
            self.update_feature_list(file_name)
            self.reset_input_fields_text()
            self.update_input_and_labels()

    def get_import(self):
        '''
        Makes a prediction based on the imported file
        '''
        self.reset_input_fields_text()
        result = 0
        import_path_null = False
        import_path = QFileDialog.getOpenFileName(filter='Excel files (*.xlsx)')[0]
        # if you open the window file explorer and click cancel
        if import_path == "":
            import_path_null = True
        sex = get_sex_int(self.combo1.currentText())
        if not import_path_null:
            result, self.sealDataImport = make_prediction(import_path, sex, self.model, self.featureList)
            seal_tag = self.get_seal_tag(import_path)
            if seal_tag == "":
                pop_message_box("Check that the correct file is being uploaded and contains a seal tag")
                result = 0
            else:
                seal_tag = "T" + seal_tag
                self.sealDataImport = np.concatenate((np.array([seal_tag]), self.sealDataImport), axis=0)
            self.sealDataManualInput = None
        if result != 0:
            self.output_label.setText("Seal tag - " + seal_tag + "\n\n" + result)

    def get_seal_tag(self, import_path):
        '''
        Retrieves the seal tag of the seal whose data is present in the file
        '''
        new_seal_data = pd.read_excel(import_path).to_numpy()
        seal_tag = np.where((new_seal_data == 'Rhb. number:') | (new_seal_data == 'Rhb. number: '))[0]
        tag = ""
        if np.where((new_seal_data == 'Rhb. number:') | (new_seal_data == 'Rhb. number: '))[0].size != 0:
            for i in range(1, 5):
                if not (pd.isnull(new_seal_data[seal_tag[0]][i])):
                    tag = new_seal_data[seal_tag[0]][i]
        return tag

    def get_manual_input_prediction(self):
        '''
        Makes a prediction based on the inputted values
        '''
        sex = get_sex_int(self.combo1.currentText())
        seal_tag = self.seal_tag_input.text()
        data = []
        try:
            for feature in self.featureList:
                data.append(float(self.featureInputDict.get(feature).text()))
            self.sealDataManualInput = data
            result, survival = find_prediction(data, self.model, sex)
            self.sealDataManualInput.append(sex)
            self.sealDataManualInput.append(survival)
            self.sealDataManualInput.insert(0, seal_tag)
            self.output_label.setText("Seal tag - " + seal_tag + "\n\n" + result)
            self.sealDataImport = None
        except ValueError:
            pop_message_box(
                "Something went wrong.\nCheck that you filled in all the boxes only entered numbers.\nMake sure to write decimals with a dot ( . ) and not a comma ( , )")

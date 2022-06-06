from openpyxl import Workbook

########################################################################################################################
# Used to create a feature file
# Run this file to create a feature file
########################################################################################################################


book = Workbook()
sheet = book.active
# Fill in the labels/tittle
sheet.cell(row=1, column=1).value = "MODEL_NAME: "
sheet.cell(row=2, column=1).value = "WBC"
sheet.cell(row=3, column=1).value = "LYMF"
sheet.cell(row=4, column=1).value = "GRAN"
sheet.cell(row=5, column=1).value = "MID"
sheet.cell(row=6, column=1).value = "RBC"
sheet.cell(row=7, column=1).value = "HGB"
sheet.cell(row=8, column=1).value = "MCH"
sheet.cell(row=9, column=1).value = "MCHC"
sheet.cell(row=10, column=1).value = "MPV"
sheet.cell(row=11, column=1).value = "PLT"

# Fill in the details for the default model
sheet.cell(row=1, column=2).value = "defaultModel"
sheet.cell(row=2, column=2).value = 1
sheet.cell(row=3, column=2).value = 1
sheet.cell(row=4, column=2).value = 1
sheet.cell(row=5, column=2).value = 1
sheet.cell(row=6, column=2).value = 1
sheet.cell(row=7, column=2).value = 1
sheet.cell(row=8, column=2).value = 1
sheet.cell(row=9, column=2).value = 1
sheet.cell(row=10, column=2).value = 1
sheet.cell(row=11, column=2).value = 1

book.save('featuresChecklist.xlsx')
